#! /workspaces/Python-and-Pandas/swap_trafic_project/venv_swap/bin/python3

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from decimal import Decimal
import sys
import os
import warnings

def prGreen(skk): print("\033[92m {}\033[00m" .format(skk))
def prRed(skk): print("\033[91m {}\033[00m" .format(skk))


warnings.filterwarnings("ignore", message="Workbook contains no default style")

# Prompt user to enter file names and check if they have .xlsx extension
WO_file = input("\nEnter the full path of the WO file: ")
if not WO_file.endswith(".xlsx"):
    WO_file += ".xlsx"
# Verify that the files exist
if not os.path.isfile(WO_file):
    raise FileNotFoundError(f"WO file '{WO_file}' not found.")

# Prompt user to enter file names and check if they have .xlsx extension
config_file = input("Enter the full path of the exported configuration file: ")
if not config_file.endswith(".xlsx"):
    config_file += ".xlsx"
# Verify that the files exist
if not os.path.isfile(config_file):
    raise FileNotFoundError(f"Configuration file '{config_file}' not found.")

# Load the replacement mapping file
WO_data = pd.read_excel(WO_file)

# Creating a dictionary for mapping from OLD values (columns B-E) to NEW values (columns G-J)
replacement_mapping = {}
replacement_counts = {}
column_mapping = {
    'OLD-IP': 'NEW-IP',
    'OLD-Mask': 'NEW-Mask',
    'OLD-VLAN': 'NEW-VLAN',
    'OLD-Gateway': 'NEW-Gateway'
}
excluded_cells = {
    "OMCH": "H3",
    "NODEBIP(RNC)": "P3"
}

# Build the replacement dictionary and initialize counts
for _, row in WO_data.iterrows():
    for old_col, new_col in column_mapping.items():
        search_value = row[old_col]
        if isinstance(search_value, float): 
            search_value = str(Decimal(search_value))
        #print (f"Search value = " , {search_value})

        replace_value = row[new_col]
        if isinstance(replace_value, float):
            replace_value = str(Decimal(replace_value))
        #print (f"Replace value = " , {replace_value})

        
        if pd.notna(search_value) and pd.notna(replace_value) and search_value != replace_value:
            replacement_mapping[search_value] = replace_value
            replacement_counts[search_value] = 0  # Initialize replacement count

# Load the configuration workbook
config_workbook = load_workbook(config_file)

# Define a green fill for replaced cells
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

# Iterate over each sheet in the config workbook and perform replacements with green fill
for sheet in config_workbook.sheetnames:
    worksheet = config_workbook[sheet]
    for row in worksheet.iter_rows():
        for cell in row:
            # Skip excluded cells
            if sheet in excluded_cells and cell.coordinate == excluded_cells[sheet]:
                #print ("Cell coordinate : ", cell.coordinate)
                continue
            # Convert cell value to string (or float) for comparison
            if cell.value is not None and str(cell.value) in replacement_mapping:
                # Increment the count for the OLD value
                replacement_counts[str(cell.value)] += 1
                # Replace cell value and apply green fill
                cell.value = replacement_mapping[str(cell.value)]
                cell.fill = green_fill

# Save the modified workbook

output_path = input("Enter the full path where you want to save the new config file : ")
os.path.exists(output_path)
if not os.path.exists(output_path):
    raise FileNotFoundError(f"The given path '{output_path}' not found.")

output_file = "New_" + os.path.basename(config_file)
config_workbook.save(output_file)
prGreen("\nNew config file saved as: " )
print (output_file)

# Display the replacement summary
prGreen("\nReplacement Summary:")
for old_value , count in replacement_counts.items():
    if old_value != "NaN":
        new_value = replacement_mapping[old_value]
        print(f"{old_value} replaced with {new_value} {count} time(s).")